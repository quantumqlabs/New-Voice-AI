from flask import Flask, render_template, request, jsonify
from flask_cors import CORS
import json
import os
from datetime import datetime, timedelta
import pandas as pd
import uuid
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path


import re

from conversation_simulator import VoiceConversationSimulator
from conversation_flows import get_conversation_flows


from elevenlabs_service import ElevenLabsTTS
from flask import send_file
import io
load_dotenv()

app = Flask(__name__)
CORS(app)

active_conversations = {}
EXCEL_FILE_PATH = "voice_prem2_conversations_log.xlsx"

elevenlabs_tts = ElevenLabsTTS()

def initialize_excel_file():
    """Initialize the Excel file with proper headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE_PATH):
        df = pd.DataFrame(columns=[
            'Conversation ID', 'Date', 'Time Start', 'Time End', 'Duration (MM:SS)', 
            'Duration (Minutes)', 'Customer Name', 'Phone Number', 'Sector', 
            'Agent Name', 'Call Status', 'Total Interactions', 'Interest Level', 
            'Lead Score (1-10)', 'Action Required', 'Next Action', 'Action Assignee',
            'Conversation Summary', 'Customer Responses Count', 'AI Responses Count',
            'Conversation Stage Reached', 'Information Gathered', 'Full Conversation Log'
        ])
        
        with pd.ExcelWriter(EXCEL_FILE_PATH, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Conversations', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Conversations']
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            column_widths = {
                'A': 38, 'B': 12, 'C': 12, 'D': 12, 'E': 15, 'F': 15,
                'G': 20, 'H': 15, 'I': 15, 'J': 15, 'K': 12, 'L': 15,
                'M': 15, 'N': 15, 'O': 15, 'P': 30, 'Q': 20, 'R': 40,
                'S': 20, 'T': 20, 'U': 30, 'V': 100
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
        
        print(f"✅ Created new Excel log file: {EXCEL_FILE_PATH}")
    else:
        print(f"✅ Using existing Excel log file: {EXCEL_FILE_PATH}")

def append_conversation_to_excel(conversation_data):
    """Append a conversation record to the Excel file"""
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            df_existing = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Conversations')
        else:
            df_existing = pd.DataFrame()
        
        new_record = pd.DataFrame([conversation_data])
        df_updated = pd.concat([df_existing, new_record], ignore_index=True)
        
        with pd.ExcelWriter(EXCEL_FILE_PATH, engine='openpyxl', mode='w') as writer:
            df_updated.to_excel(writer, sheet_name='Conversations', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Conversations']
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            column_widths = {
                'A': 38, 'B': 12, 'C': 12, 'D': 12, 'E': 15, 'F': 15,
                'G': 20, 'H': 15, 'I': 15, 'J': 15, 'K': 12, 'L': 15,
                'M': 15, 'N': 15, 'O': 15, 'P': 30, 'Q': 20, 'R': 40,
                'S': 20, 'T': 20, 'U': 30, 'V': 100
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        print(f"✅ Conversation saved to Excel: {conversation_data['Customer Name']} - {conversation_data['Sector']}")
        return True
        
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")
        return False

def is_off_topic_question(response):
    """Check if the response is off-topic - AI will handle it but NOT end conversation"""
    response_lower = response.lower().strip()
    
    # Identity questions - AI should answer these
    identity_questions = [
        'what is your name', 'what\'s your name', 'whats your name',
        'who are you', 'where are you from', 'where you from',
        'which company', 'what company', 'your name', 'tell me your name'
    ]
    
    if any(q in response_lower for q in identity_questions):
        return True, 'identity'
    
    # CRITICAL: Exclude business/financial contexts FIRST
    business_terms = [
        'loan', 'lakh', 'lakhs', 'crore', 'crores', 'rupees', 'rs', 'inr',
        'amount', 'expecting', 'need', 'require', 'borrow', 'budget',
        'salary', 'income', 'emi', 'interest', 'rate', 'credit', 'payment',
        'bhk', 'property', 'apartment', 'flat', 'house', 'square feet', 'sqft',
        'checkup', 'consultation', 'appointment', 'health', 'medical', 'insurance',
        'application', 'eligible', 'document', 'apply', 'process'
    ]
    
    # If any business term is present, this is ON-TOPIC
    if any(term in response_lower for term in business_terms):
        print(f"[OFF-TOPIC] ✅ Business context - ON-TOPIC")
        return False, None
    
    # Math questions - only if NO business context
    math_indicators = ['what is', 'what\'s', 'whats', 'calculate', 'equals']
    has_math_indicator = any(indicator in response_lower for indicator in math_indicators)
    has_numbers = any(char.isdigit() for char in response)
    has_operator = any(op in response for op in ['+', '-', '*', '/', 'x', '×', '÷'])
    
    # Pure math: "what is 2+2" or "5*5"
    if (has_math_indicator and has_numbers and has_operator):
        print(f"[OFF-TOPIC] 🔢 Pure math question detected")
        return True, 'math'
    
    # Short calculation with operators
    if has_numbers and has_operator and len(response.split()) <= 5:
        print(f"[OFF-TOPIC] 🔢 Simple calculation detected")
        return True, 'math'
    
    # General knowledge questions - but NOT service-related
    general_questions = [
        'who is', 'who was', 'what is the capital', 'when did', 'where is', 
        'how tall', 'what year', 'when was', 'who invented', 'what color',
        'which country', 'what happened', 'who won', 'how many'
    ]
    
    if any(q in response_lower for q in general_questions):
        service_check_terms = ['loan', 'credit', 'bank', 'property', 'house', 'medical', 'health', 'checkup', 'apartment']
        if not any(term in response_lower for term in service_check_terms):
            print(f"[OFF-TOPIC] 📚 General knowledge question")
            return True, 'general'
    
    print(f"[OFF-TOPIC] ✅ ON-TOPIC response")
    return False, None

@app.route('/')
def index():
    """Serve the main HTML page"""
    return render_template('index.html')

@app.route('/api/text-to-speech', methods=['POST'])
def text_to_speech():
    """Generate speech audio using ElevenLabs"""
    try:
        data = request.get_json()
        text = data.get('text')
        
        if not text:
            return jsonify({'success': False, 'error': 'No text provided'}), 400
        
        audio_data = elevenlabs_tts.text_to_speech(text)
        
        if audio_data:
            return send_file(
                io.BytesIO(audio_data),
                mimetype='audio/mpeg',
                as_attachment=False
            )
        else:
            return jsonify({
                'success': False, 
                'error': 'ElevenLabs not configured or failed',
                'fallback': True
            }), 200
            
    except Exception as e:
        print(f"Error in text_to_speech: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/start_conversation', methods=['POST'])
def start_conversation():
    """Initialize a new conversation session"""
    try:
        data = request.get_json()
        customer_name = data.get('customerName')
        phone_number = data.get('phoneNumber') 
        sector = data.get('sector')
        
        if not all([customer_name, phone_number, sector]):
            return jsonify({
                'success': False,
                'error': 'Missing required fields'
            }), 400
        
        if not os.getenv('OPENAI_API_KEY'):
            return jsonify({
                'success': False,
                'error': 'No OpenAI API key configured. Please set OPENAI_API_KEY in your .env file.'
            }), 500
        
        conversation_id = str(uuid.uuid4())
        simulator = VoiceConversationSimulator(customer_name, phone_number, sector)
        active_conversations[conversation_id] = simulator
        
        try:
            opening_message = simulator.get_opening_message()
        except Exception as e:
            print(f"Error generating AI opening message: {e}")
            opening_message = f"Hello {customer_name}! This is a call regarding our {sector} services. Do you have a moment to speak?"
        
        print(f"[START] Customer: {customer_name}, Sector: {sector}")
        print(f"[START] Opening: {opening_message}")
        
        return jsonify({
            'success': True,
            'conversation_id': conversation_id,
            'opening_message': opening_message,
            'customer_info': {
                'name': customer_name,
                'phone': phone_number,
                'sector': sector.replace('_', ' ').title()
            }
        })
        
    except Exception as e:
        print(f"Error in start_conversation: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/process_response', methods=['POST'])
def process_response():
    """Process customer response - AI handles off-topic then returns to main flow"""
    try:
        data = request.get_json()
        conversation_id = data.get('conversation_id')
        customer_response = data.get('customer_response')
        
        if conversation_id not in active_conversations:
            return jsonify({
                'success': False,
                'error': 'Conversation not found'
            }), 404
        
        simulator = active_conversations[conversation_id]
        
        print(f"[PROCESS] Customer: {customer_response}")
        
        # Check if off-topic (but DON'T end conversation - just let AI handle it)
        is_off_topic, question_type = is_off_topic_question(customer_response)
        
        if is_off_topic:
            print(f"[PROCESS] Off-topic {question_type} - AI will answer and redirect")
            # Don't use separate handler - let AI service handle it naturally
            # It will answer the question and smoothly redirect back to main topic
        
        # Process response normally - AI will handle everything
        try:
            ai_response = simulator.get_next_ai_response(customer_response)
            
            if simulator.closing_sent or ai_response is None:
                print(f"[PROCESS] ✅ Conversation ended")
                
                if ai_response is None:
                    ai_response = get_conversation_flows()[simulator.sector]['closing']
                
                return jsonify({
                    'success': True,
                    'ai_response': ai_response,
                    'conversation_ended': True,
                    'conversation_state': {
                        'interest_level': simulator.customer_interest_level,
                        'lead_score': simulator.lead_score,
                        'next_action': simulator.next_action,
                        'duration': calculate_duration(simulator.start_time, simulator.end_time),
                        'total_interactions': simulator.total_interactions
                    }
                })
        except Exception as e:
            print(f"Error generating AI response: {e}")
            ai_response = get_conversation_flows()[simulator.sector]['closing']
            simulator.end_time = datetime.now()
            simulator.closing_sent = True
            simulator.set_final_actions()
        
        duration = calculate_duration(simulator.start_time, simulator.end_time)
        
        print(f"[PROCESS] AI: {ai_response[:80]}...")
        print(f"[PROCESS] Ended: {simulator.closing_sent}")
        
        return jsonify({
            'success': True,
            'ai_response': ai_response,
            'conversation_ended': simulator.closing_sent,
            'conversation_state': {
                'interest_level': simulator.customer_interest_level,
                'lead_score': simulator.lead_score,
                'next_action': simulator.next_action,
                'duration': duration,
                'total_interactions': simulator.total_interactions,
                'current_stage': getattr(simulator, 'conversation_state', 'unknown')
            }
        })
        
    except Exception as e:
        print(f"Error in process_response: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/end_conversation', methods=['POST'])
def end_conversation():
    """End conversation and get comprehensive results"""
    try:
        data = request.get_json()
        conversation_id = data.get('conversation_id')
        
        if conversation_id not in active_conversations:
            return jsonify({
                'success': False,
                'message': 'Conversation not found'
            }), 404
        
        simulator = active_conversations[conversation_id]
        
        if not simulator.end_time:
            simulator.end_time = datetime.now()
        if not simulator.closing_sent:
            closing_message = get_conversation_flows()[simulator.sector]['closing']
            simulator.conversation_log.append(f"AI Agent: {closing_message}")
            simulator.closing_sent = True
            simulator.set_final_actions()
        
        duration = calculate_duration(simulator.start_time, simulator.end_time)
        
        results = {
            'customer_name': simulator.customer_name,
            'phone_number': simulator.phone_number,
            'sector': simulator.sector.replace('_', ' ').title(),
            'duration': duration,
            'interest_level': simulator.customer_interest_level,
            'lead_score': simulator.lead_score,
            'next_action': simulator.next_action,
            'action_assignee': simulator.action_assignee,
            'action_required': simulator.action_required,
            'call_status': 'Completed',
            'conversation_log': simulator.conversation_log,
            'remarks': simulator.remarks,
            'start_time': simulator.start_time.strftime("%H:%M:%S"),
            'end_time': simulator.end_time.strftime("%H:%M:%S"),
        }
        
        conversation_data = {
            'Conversation ID': conversation_id,
            'Date': datetime.now().strftime("%Y-%m-%d"),
            'Time Start': simulator.start_time.strftime("%H:%M:%S"),
            'Time End': simulator.end_time.strftime("%H:%M:%S"),
            'Duration (MM:SS)': duration,
            'Duration (Minutes)': round((simulator.end_time - simulator.start_time).total_seconds() / 60, 2),
            'Customer Name': simulator.customer_name,
            'Phone Number': simulator.phone_number,
            'Sector': simulator.sector,
            'Agent Name': simulator.ai_service.agent_personas[simulator.sector]['name'],
            'Call Status': 'Completed',
            'Total Interactions': simulator.total_interactions,
            'Interest Level': simulator.customer_interest_level,
            'Lead Score (1-10)': simulator.lead_score,
            'Action Required': simulator.action_required,
            'Next Action': simulator.next_action,
            'Action Assignee': simulator.action_assignee,
            'Conversation Summary': simulator.remarks,
            'Customer Responses Count': len([l for l in simulator.conversation_log if l.startswith('Customer:')]),
            'AI Responses Count': len([l for l in simulator.conversation_log if l.startswith('AI Agent:')]),
            'Conversation Stage Reached': simulator.conversation_state,
            'Information Gathered': simulator.customer_preference or 'N/A',
            'Full Conversation Log': '\n'.join(simulator.conversation_log)
        }

        # ✅ Save to Excel
        append_conversation_to_excel(conversation_data)
        print(f"✅ Conversation saved to Excel for {simulator.customer_name}")

        return jsonify({
            'success': True,
            'results': results
        })

    except Exception as e:
        print(f"❌ Error in end_conversation: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

def save_conversation_to_excel(conversation_id, simulator):
    """Save conversation data to Excel file"""
    try:
        duration_seconds = (simulator.end_time - simulator.start_time).total_seconds() if simulator.end_time else 0
        duration_formatted = f"{int(duration_seconds // 60)}:{int(duration_seconds % 60):02d}"
        
        conversation_text = '\n\n'.join(simulator.conversation_log)
        
        # Format information gathered
        info_gathered = json.dumps(simulator.information_gathered, indent=2)
        
        conversation_data = {
            'Conversation ID': conversation_id,
            'Date': simulator.start_time.strftime('%Y-%m-%d'),
            'Time Start': simulator.start_time.strftime('%H:%M:%S'),
            'Time End': simulator.end_time.strftime('%H:%M:%S') if simulator.end_time else '',
            'Duration (MM:SS)': duration_formatted,
            'Duration (Minutes)': round(duration_seconds / 60, 1),
            'Customer Name': simulator.customer_name,
            'Phone Number': simulator.phone_number,
            'Sector': simulator.sector.replace('_', ' ').title(),
            'Agent Name': simulator.ai_service.agent_personas[simulator.sector]['name'],
            'Call Status': 'Completed',
            'Total Interactions': simulator.total_interactions,
            'Interest Level': simulator.customer_interest_level,
            'Lead Score (1-10)': simulator.lead_score,
            'Action Required': simulator.action_required,
            'Next Action': simulator.next_action,
            'Action Assignee': simulator.action_assignee,
            'Conversation Summary': simulator.remarks,
            'Customer Responses Count': len([msg for msg in simulator.conversation_log if msg.startswith("Customer:")]),
            'AI Responses Count': len([msg for msg in simulator.conversation_log if msg.startswith("AI Agent:")]),
            'Conversation Stage Reached': simulator.current_stage,
            'Information Gathered': info_gathered,
            'Full Conversation Log': conversation_text
        }
        
        success = append_conversation_to_excel(conversation_data)
        
        if success:
            if conversation_id in active_conversations:
                del active_conversations[conversation_id]
                print(f"✅ Cleaned up conversation {conversation_id}")
        
        return success
        
    except Exception as e:
        print(f"Error in save_conversation_to_excel: {e}")
        return False

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    ai_status = "Available" if os.getenv('OPENAI_API_KEY') else "Not Configured"
    
    return jsonify({
        'status': 'healthy',
        'active_conversations': len(active_conversations),
        'ai_service': ai_status,
        'excel_file': EXCEL_FILE_PATH,
        'excel_exists': os.path.exists(EXCEL_FILE_PATH),
        'features': [
            'Structured Banking Flow (Eligibility → Process → Meeting)',
            'Smart Off-Topic Handling (Answers then Redirects)',
            'Meeting/Appointment Scheduling',
            'Information Extraction',
            'Stage-Based Conversation',
            'Natural Flow Like Medical Sector'
        ]
    })

def calculate_duration(start_time, end_time):
    """Calculate conversation duration"""
    if not end_time:
        return "00:00"
    
    duration_seconds = (end_time - start_time).total_seconds()
    minutes = int(duration_seconds // 60)
    seconds = int(duration_seconds % 60)
    return f"{minutes:02d}:{seconds:02d}"



if __name__ == '__main__':
    initialize_excel_file()
    
    api_key = os.getenv('OPENAI_API_KEY')
    if api_key:
        print("✅ OpenAI API configured")
    else:
        print("❌ No OpenAI API key - set OPENAI_API_KEY in .env")
        print()
    
    os.makedirs('templates', exist_ok=True)
    
    print("🚀 AI Voice Conversation Simulator - STRUCTURED FLOW")
    print("✨ Features:")
    print("   ✓ Banking: Eligibility → Process → Meeting Scheduling")
    print("   ✓ Real Estate: Preference → Budget → Site Visit")
    print("   ✓ Medical: Service → Details → Appointment")
    print("   ✓ Smart Off-Topic Handling (Answer & Redirect)")
    print("   ✓ Information Extraction & Tracking")
    print(f"📊 Excel: {os.path.abspath(EXCEL_FILE_PATH)}")
    print("🌐 URL: http://localhost:5000")
    print()
    
    app.run(debug=False, host='0.0.0.0', port=5000)